import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()



import openpyxl, sys
from synprodutos.models import *
from django.db.models import Count

wb = openpyxl.load_workbook('SynProdutos.xlsx')
sheet = wb['Produtos_Onix']

#populando syncomercial.Produto_Onix
for rowNum in range(2, sheet.max_row + 1):
    prod = Produto_Onix()
    prod.id = sheet.cell(column = 1, row = rowNum).value
    prod.descricao = sheet.cell(column = 2, row = rowNum).value
    prod.prefixo = sheet.cell(column = 3, row = rowNum).value
    prod.embalagem = sheet.cell(column = 4, row = rowNum).value
    prod.unidade = sheet.cell(column = 5, row = rowNum).value
    prod.volume = sheet.cell(column = 6, row = rowNum).value
    prod.status = sheet.cell(column = 7, row = rowNum).value
    prod.criado_em = django.utils.timezone.now()
    prod.atualizado_em = django.utils.timezone.now()
    prod.save()
    del prod

print('Done!')