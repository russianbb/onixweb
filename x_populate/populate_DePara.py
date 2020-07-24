import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()



import openpyxl, sys
from synprodutos.models import *
from django.db.models import Count

wb = openpyxl.load_workbook('SynProdutos_DePara.xlsx')
sheet = wb['Sheet1']

#populando syncomercial.Produto_Distribuidor
for rowNum in range(2, sheet.max_row + 1):
    prod = Produto_Distribuidor()
    prod.produto_onix_id = sheet.cell(column = 1, row = rowNum).value
    prod.distribuidor_id = sheet.cell(column = 2, row = rowNum).value
    prod.codigo = sheet.cell(column = 3, row = rowNum).value
    prod.descricao = sheet.cell(column = 4, row = rowNum).value
    prod.criado_em = django.utils.timezone.now()
    prod.atualizado_em = django.utils.timezone.now()
    prod.save()
    del prod
    print(rowNum)

print('Done!')
