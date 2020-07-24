import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()



import openpyxl, sys
from syncomercial.models import *
from django.db.models import Count

wb = openpyxl.load_workbook('SynComercial_Filial.xlsx')
sheet = wb['Sheet1']

#populando syncomercial.Produto_Onix
for rowNum in range(749, sheet.max_row + 1):
    filial = Filial()
    filial.id = sheet.cell(column = 1, row = rowNum).value
    filial.codigo = sheet.cell(column = 2, row = rowNum).value
    filial.nome = sheet.cell(column = 3, row = rowNum).value
    filial.cnpj = sheet.cell(column = 4, row = rowNum).value
    filial.responsavel = sheet.cell(column = 5, row = rowNum).value
    filial.email = sheet.cell(column = 6, row = rowNum).value
    filial.ddd = sheet.cell(column = 7, row = rowNum).value
    filial.telefone = sheet.cell(column = 8, row = rowNum).value
    filial.endereco = sheet.cell(column = 9, row = rowNum).value
    filial.cidade = sheet.cell(column = 10, row = rowNum).value
    filial.estado = sheet.cell(column = 11, row = rowNum).value
    filial.codibge = sheet.cell(column = 12, row = rowNum).value
    filial.latitude = sheet.cell(column = 13, row = rowNum).value
    filial.longitude = sheet.cell(column = 14, row = rowNum).value
    filial.controla_estoque = sheet.cell(column = 15, row = rowNum).value
    filial.criado_em = sheet.cell(column = 16, row = rowNum).value
    filial.atualizado_em = sheet.cell(column = 17, row = rowNum).value
    filial.status = sheet.cell(column = 18, row = rowNum).value
    filial.distribuidor_id = sheet.cell(column = 19, row = rowNum).value
    filial.save()
    print(f'Row:{rowNum} - ID: {filial.id}')
    del filial

print('Done!')