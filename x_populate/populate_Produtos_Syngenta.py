import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()



import openpyxl, sys
from synprodutos.models import *
from django.db.models import Count

wb = openpyxl.load_workbook('SynProdutos.xlsx')
sheet = wb['Produtos_Syngenta']

#populando syncomercial.Produto_Onix
for rowNum in range(2, sheet.max_row + 1):
    prod = Produto_Syngenta()
    prod.produto_onix_id = sheet.cell(column = 1, row = rowNum).value
    prod.agicode = sheet.cell(column = 2, row = rowNum).value
    prod.familia = sheet.cell(column = 3, row = rowNum).value
    prod.descricao = sheet.cell(column = 4, row = rowNum).value
    prod.criado_em = django.utils.timezone.now()
    prod.atualizado_em = django.utils.timezone.now()
    prod.save()
    del prod

print('Done!')