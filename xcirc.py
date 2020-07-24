import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()



import openpyxl, sys
from syncomercial.models import *
from django.db.models import Count

wb = openpyxl.load_workbook('MODELO_IMPORTACAO_TESTE.xlsx')
base = wb['BASE']

#VALIDACAO 1 --> Valida o CODE SAP
codeSap = []
for rowNum in range(2, base.max_row + 1):
    #Appenda todos os CODE SAP preenchidos na lista
    codeSap.append(base.cell(column = 1, row = rowNum).value)

#Remove as duplicatas
codeSap = list(dict.fromkeys(codeSap))
if len(codeSap) > 1:
    # Mudar abaixo para retornar erro via javascript TODO
    print('Mais de um CODE SAP cadastrado.')
    sys.exit()
elif codeSap[0] == None :
    # Mudar abaixo para retornar erro via javascript TODO
    print('CODE SAP nao preenchido.')
    sys.exit()

distribuidor = Distribuidor.objects.get(code_sap=codeSap[0])
#Estudar Try-Exception para query acima. Se nao encontra, crasha o codigo TODO
if not distribuidor:
    # Mudar abaixo para retornar erro via javascript TODO
    print(f'CODE SAP: {codeSap[0]} nao Cadastrado, favor verificar cadastros.')
    sys.exit()


#VALIDACAO 2 --> Valida cadastro de filiais

#Busca filiais do distribuidor no banco de dados
filiais = Filial.objects.filter(distribuidor=distribuidor.id)

#Verifica cada linha, buscando por unico match de filial cadastrada
for rowNum in range(2, 5):
    cnpj = base.cell(column = 8, row = rowNum).value
    codigo = base.cell(column = 6, row = rowNum).value
    nome = base.cell(column = 7, row = rowNum).value
    print(f'\nLinha: {rowNum} - Cod: {codigo} -- Nome: {nome}')
    
    if len(filiais.filter(cnpj=cnpj)) == 0:
        #Se CNPJ nao estiver cadastrado no BDD
        # Mudar abaixo para retornar erro via javascript TODO
        print('Nao cadastrado: CNPJ')
    
    if len(filiais.filter(cnpj=cnpj)) == 1:
        #Se CNPJ tiver apenas 1 cadastro no BDD
        # Cadastro ok, prosseguir
        print('Cadastro ok: CNPJ')
        filial = filiais.get(cnpj=cnpj)
        print(f'ID: {filial.id}')
        
        
    if len(filiais.filter(cnpj=cnpj)) > 1:
        #Se CNPJ tiver mais de 1 cadastro no BDD
        # Comeca a buscar mais a fundo, considerando CODIGO da filial
        
        if len(filiais.filter(cnpj=cnpj, codigo=codigo)) == 0:
            #Se CNPJ nao estiver cadastrado no BDD
            # Mudar abaixo para retornar erro via javascript TODO
            print('Nao cadastrado: CNPJ, CODIGO')
            
        
        if len(filiais.filter(cnpj=cnpj, codigo=codigo)) == 1:
            #Se CNPJ tiver apenas 1 cadastro no BDD
            # Cadastro ok, prosseguir
            print('Cadastro ok: CNPJ, CODIGO')
            filial = filiais.get(cnpj=cnpj, codigo=codigo)
            print(f'ID: {filial.id}')
            
        
        if len(filiais.filter(cnpj=cnpj, codigo=codigo)) > 1:
            #Se CNPJ tiver mais de 1 cadastro no BDD
            # Comeca a buscar mais a fundo, considerando NOME da filial
            
            if len(filiais.filter(cnpj=cnpj, codigo=codigo, nome=nome)) == 0:
                #Se CNPJ nao estiver cadastrado no BDD
                # Mudar abaixo para retornar erro via javascript TODO
                print('Nao cadastrado: CNPJ, CODIGO, NOME')
                
            
            if len(filiais.filter(cnpj=cnpj, codigo=codigo, nome=nome)) == 1:
                #Se CNPJ tiver apenas 1 cadastro no BDD
                # Cadastro ok, prosseguir
                print('Cadastro ok: CNPJ, CODIGO, NOME')
                filial = filiais.get(cnpj=cnpj, codigo=codigo, nome=nome)
                print(f'ID: {filial.id}')
            
            else: 
                print('Fudeu, n faco ideia de pq parou aqui....')


